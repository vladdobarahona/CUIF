# -*- coding: utf-8 -*-
"""
Aplicación Streamlit – Información Financiera CUIF (SFC)
-------------------------------------------------------

Created on Thu Jan 15 15:10:38 2026
Author: Vladimir Alonso Barahona Palacios

Descripción:
------------
Aplicación interactiva desarrollada en Streamlit para la descarga,
validación y procesamiento de información financiera con fines de supervisión,
correspondiente al Catálogo Único de Información Financiera – CUIF
publicado por la Superintendencia Financiera de Colombia (SFC).

La aplicación permite:
- Consultar la fecha máxima disponible del dataset.
- Descargar información entre un rango de fechas.
- Validar la cantidad de registros antes de descargar.
- Procesar, limpiar y consolidar información CUIF por moneda (Total).
- Integrar los datos con una plantilla de cuentas NIIF.
- Generar un archivo Excel con estructura predefinida.
- Permitir la descarga directa del archivo consolidado.

Fuente de datos:
----------------
Datos abiertos – Superintendencia Financiera de Colombia:
https://www.superfinanciera.gov.co

Repositorio oficial del dataset:
https://www.datos.gov.co/Hacienda-y-Cr-dito-P-blico/Informaci-n-financiera-con-fines-de-supervisi-n-CU/mxk5-ce6w

Categoría:
----------
Hacienda y Crédito Público

Notas:
------
- Los datos son consultados en tiempo real desde la API de datos.gov.co (Socrata).
- La aplicación soporta descargas masivas mediante paginación.
- Exporta el reporte final en formato Excel compatible con reportes regulatorios.

"""

import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re


# ==============================
# Estilos personalizados
# ==============================
st.markdown("""
<style>
    /* Fondo de toda la aplicación */
    .stApp {
        background: #ffffff !important;
        font-family: "Segoe UI", "Frutiger", "Helvetica Neue", sans-serif;
        padding-top: 20px;
    }

    /* Título principal */
    .main-title {
        color: rgb(120,154,61);
        font-size: 2.5rem;
        font-weight: 700;
        line-height: 1.25;
        margin-top: 15px;
        margin-bottom: 0px;
    }

    /* Subtítulo */
    .sub-title {
        color: #4a4a4a;
        font-size: 1.1rem;
        margin-top: -5px;
        margin-bottom: 25px;
    }

    /* Fondo general de la página (fuera del contenedor blanco) */
    body {
        background-color: rgb(171,190,76) !important;
    }
</style>
""", unsafe_allow_html=True)


# ==============================
# LOGO + TÍTULO
# ==============================
col1, col2 = st.columns([1, 3])

with col1:
    st.image(
        "https://www.finagro.com.co/sites/default/files/logo-front-finagro.png",
        width=180
    )

with col2:
    st.markdown(
        """
        <h1 class="main-title">
            Información Financiera CUIF – Consulta, Descarga y Procesamiento
        </h1>
        <div class="sub-title">
            Sistema de apoyo para entidades vigiladas – Moneda Total
        </div>
        """,
        unsafe_allow_html=True
    )


# -------------------------
# CONFIG
# -------------------------
BASE_URL = "https://www.datos.gov.co/resource/mxk5-ce6w.json"
LIMIT = 50000  # máximo permitido por Socrata


# -------------------------
# UNIDADES DE VALORES
# -------------------------
UNIDADES_VALOR = {
    "Sin Unidades": (1, "Pesos"),
    "Miles": (1_000, "Miles de Pesos"),
    "Millones": (1_000_000, "Millones de Pesos"),
    "Cientos de Millones": (100_000_000, "Cientos de Millones de Pesos"),
    "Miles de Millones": (1_000_000_000, "Miles de Millones de Pesos"),
    "Billones": (1_000_000_000_000, "Billones de Pesos"),
}

# -------------------------
# FUNCIONES API
# -------------------------
def max_fecha():
    """Consulta la fecha máxima disponible en el dataset."""
    query = '''
    https://www.datos.gov.co/resource/mxk5-ce6w.json?$query=
    SELECT max(fecha_corte)
    '''
    r = requests.get(query)
    if r.status_code != 200:
        return None
    data = r.json()
    if data:
        return data[0]["max_fecha_corte"]
    return None


def conteo(tipo_entidad: str, fecha_desde: str, fecha_hasta: str) -> int:
    """
    Devuelve el número de registros en el dataset CUIF para:
    - tipo_entidad (str exacto como aparece en la SFC)
    - rango de fechas [fecha_desde, fecha_hasta] en formato YYYY-MM-DD

    Retorna: int (count)
    Lanza: Exception si la API responde con error distinto de 200.
    """
    where = (
        f"fecha_corte between '{fecha_desde}T00:00:00' and '{fecha_hasta}T23:59:59'"
        f" AND nombre_moneda = 'Total'"
        f" AND nombre_tipo_entidad = '{tipo_entidad}'"
    )

    params = {
        "$select": "count(*)",
        "$where": where
    }

    r = requests.get(BASE_URL, params=params)
    if r.status_code != 200:
        raise Exception(f"Error HTTP {r.status_code}: {r.text}")

    data = r.json()
    # Respuesta típica: [{'count': '12345'}]
    try:
        return int(data[0].get("count", 0))
    except (IndexError, ValueError, TypeError):
        return 0

def descargar_datos(tipo_entidad,fecha_desde, fecha_hasta):
    """Descarga datos con paginación."""
    where_clause = (
        f"fecha_corte between '{fecha_desde}T00:00:00' and "
        f"'{fecha_hasta}T23:59:59' AND "
        f"nombre_moneda = 'Total' AND "
        f"nombre_tipo_entidad = '{tipo_entidad}'"
    )

    offset = 0
    all_rows = []

    while True:
        params = {
            "$limit": LIMIT,
            "$offset": offset,
            "$where": where_clause,
        }
        r = requests.get(BASE_URL, params=params)

        if r.status_code != 200:
            raise Exception(f"Error HTTP {r.status_code}: {r.text}")

        data = r.json()
        if not data:
            break

        all_rows.extend(data)
        offset += LIMIT

    return pd.DataFrame(all_rows)


# -------------------------
# PROCESAMIENTO
# -------------------------
def procesar_dataframe(df, plantilla_path, factor_rango):
    df_final = df.copy()
    mask = df_final['cuenta'].isin(cambio_signo_cuenta)
    df_final.loc[mask, 'valor'] = df_final.loc[mask, 'valor'].astype(str).str.replace('-', '', regex=False)
    Rango_de_Valores = factor_rango   
    df_final["valor_Rango_de_Valores"] = (
        pd.to_numeric(df_final["valor"], errors="coerce") / Rango_de_Valores
    ).round(0)

    df_final["Label"] = df_final["codigo_entidad"].astype(str) + " - " + df_final["nombre_entidad"]

    # ----- PIVOT -----
    pivot_df = pd.pivot(
        df_final,
        index=["cuenta", "nombre_cuenta"],
        columns="Label",
        values="valor_Rango_de_Valores",
    )

    pivot_df.columns.name = None
    pivot_df = pivot_df.reset_index()

    # ----- ORDENAR COLUMNAS -----
    def sort_by_prefix(col):
        match = re.match(r'^(\d+)\s*-', str(col))
        return int(match.group(1)) if match else -1

    fixed = ["cuenta", "nombre_cuenta"]
    other_cols = [c for c in pivot_df.columns if c not in fixed]
    pivot_df = pivot_df[fixed + sorted(other_cols, key=sort_by_prefix)]

    # ----- MERGE CON PLANTILLA -----
    plantilla = pd.read_excel(plantilla_path, sheet_name="Cuentas")
    plantilla = plantilla[["Cuenta", "Descripción_Cuenta"]]
    plantilla["Cuenta"] = plantilla["Cuenta"].astype(str)
    pivot_df["cuenta"] = pivot_df["cuenta"].astype(str)
    pivot_df = pivot_df.drop(columns = 'nombre_cuenta')
    plantilla = plantilla.rename(columns={"Cuenta": "cuenta", "Descripción_Cuenta": "nombre_cuenta"})
    plantilla = plantilla.set_index(["cuenta"])
    pivot_df = pivot_df.set_index(["cuenta"])
    pivot_full = plantilla.join(pivot_df, how="left").fillna(0)
    pivot_full = pivot_full.reset_index()
    return pivot_full


# -------------------------
# EXPORTAR A EXCEL
# -------------------------
def generar_excel(pivot_df, tipo_entidad, fecha_desde, factor_rango, etiqueta_unidad):
    f_informe = datetime.strptime(fecha_desde, "%Y-%m-%d").strftime("%d/%m/%Y")
    tipo_entidad_str = str(tipo_entidad)
    date_obj = datetime.strptime(fecha_desde, "%Y-%m-%d")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{nombre_hoja[tipo_entidad]}{date_obj.strftime('%m%Y')}g1m0cie"

    # Encabezados
    ws["A2"] = "Tipo de Entidad:"
    ws["B2"] = tipo_entidad_str

    ws["A3"] = "Fecha de Informe:"
    ws["B3"] = f_informe

    ws["A4"] = "Moneda:"
    ws["B4"] = "0 Total"

    ws["A5"] = "Rango de Valores:"
    ws["B5"] = str(factor_rango)
    ws["C5"] = etiqueta_unidad


    start_row = 9
    for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=False, header=True)):
        for c_idx, value in enumerate(row):
            col_letter = get_column_letter(1 + c_idx)
            ws[f"{col_letter}{start_row + r_idx}"] = value

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# -------------------------
# UI STREAMLIT
# -------------------------

st.write("Aplicación para descarga, validación y generación de reportes CUIF.")

# ---- CONSULTAR FECHA MÁXIMA -----
if st.button("Consultar Fecha Máxima Disponible"):
    fecha = max_fecha()
    if fecha:
        st.success(f"📅 Fecha máxima encontrada: **{fecha}**")
    else:
        st.error("No se pudo obtener la fecha máxima.")


# ---- DESCARGA PERSONALIZADA -----
st.subheader("Descargar datos por rango de fechas")

fecha_desde = st.date_input("Fecha Desde")
fecha_hasta = st.date_input("Fecha Hasta")

st.subheader("Seleccionar el tipo de entidad")

lista_tipo_entidad = [
    "ESTABLECIMIENTOS BANCARIOS",                                          
    "COMPANIAS DE SEGUROS GENERALES",
    "COMPANIAS DE SEGUROS DE VIDA",
    "SOCIEDADES FIDUCIARIAS",
    "COMPANIAS DE FINANCIAMIENTO COMERCI",
    "COMISIONISTAS  DE BOLSA DE VALORES",
    "INSTITUCIONES OFICIALES ESPECIALES",
    "COOPERATIVAS CARACTER FINANCIERO",
    "SISTEMAS DE PAGO DE BAJO VALOR",
    "SOCIEDADES COMISIONISTAS DE BOLSAS AGROPECUARIAS",
    "CORPORACIONES FINA11NCIERAS",
    "SOC ESP DEPOSITOS Y PAGOS ELECTRONI",
    "SOC. ADM. FONDOS DE PENSIONES Y CES",
    "ALMACENES GENERALES DE DEPOSITO",
    "SOCIEDADES COOPERATIVAS DE SEGUROS",
    "SOC ADM SIST DE  NEG Y REG SOBRE VA",
    "SOC ADM SIST DE NEG Y REG DE DIVISA",
    "PROVEEDOR DE PRECIOS PARA VALORACIO",
    "SOCIEDADES CALIFICADORAS DE VALORES",
    "SOCIEDADES DE CAPITALIZACION",
    "BOLSAS DE VALORES",
    "SOCIEDADES ADMINISTRADORAS DE DEPOSITOS CENTRALIZADOS V",
    "ENTID ADM. REGIMEN SOL PRIMA MEDIA",
    "CAMARA RIESGO CENTRAL D CONTRAPARTE",
    "BOLSAS AGROPECUARIAS",
    "ORGANISMOS DE AUTORREGULACION",
]

nombre_hoja = {
    "ESTABLECIMIENTOS BANCARIOS":'0001',                                          
    "COMPANIAS DE SEGUROS GENERALES":'0013',
    "COMPANIAS DE SEGUROS DE VIDA":'0014',
    "SOCIEDADES FIDUCIARIAS":'0005',
    "COMPANIAS DE FINANCIAMIENTO COMERCI":'0004',
    "COMISIONISTAS  DE BOLSA DE VALORES":'0085',
    "INSTITUCIONES OFICIALES ESPECIALES":'0022',
    "COOPERATIVAS CARACTER FINANCIERO": '0032',
    "SISTEMAS DE PAGO DE BAJO VALOR":'0012',
    "SOCIEDADES COMISIONISTAS DE BOLSAS AGROPECUARIAS":'0401',
    "CORPORACIONES FINA11NCIERAS": '0002',
    "SOC ESP DEPOSITOS Y PAGOS ELECTRONI":'0128',
    "SOC. ADM. FONDOS DE PENSIONES Y CES":'0023',
    "ALMACENES GENERALES DE DEPOSITO":'0006',
    "SOCIEDADES COOPERATIVAS DE SEGUROS":'0015',
    "SOC ADM SIST DE  NEG Y REG SOBRE VA":'0502',
    "SOC ADM SIST DE NEG Y REG DE DIVISA":'0501',
    "PROVEEDOR DE PRECIOS PARA VALORACIO":'0509',
    "SOCIEDADES CALIFICADORAS DE VALORES":'0084',
    "SOCIEDADES DE CAPITALIZACION":'0010',
    "BOLSAS DE VALORES":'0082',
    "SOCIEDADES ADMINISTRADORAS DE DEPOSITOS CENTRALIZADOS V":'0083',
    "ENTID ADM. REGIMEN SOL PRIMA MEDIA":'0025',
    "CAMARA RIESGO CENTRAL D CONTRAPARTE":'0504',
    "BOLSAS AGROPECUARIAS":'0400',
    "ORGANISMOS DE AUTORREGULACION":'0081',
}

cambio_signo_cuenta = ['392000']

tipo_entidad = st.selectbox(
    label="Seleccione el tipo de entidad (SFC):",
    options=lista_tipo_entidad,
    index=0,  # Default selection (0 = first item)
    placeholder="Seleccione el tipo de entidad (SFC)..."  # Optional placeholder
)

# Display the selected value
st.write(f"Has seleccionado: **{tipo_entidad}**")


st.subheader("Seleccione el Rango de Valores")

opcion_rango = st.selectbox(
    "Unidad de salida:",
    list(UNIDADES_VALOR.keys())
)

factor_rango, etiqueta_unidad = UNIDADES_VALOR[opcion_rango]

st.caption(f"Los valores se dividirán por **{factor_rango:,}** ({etiqueta_unidad})")

plantilla_file = st.file_uploader("Suba la plantilla de cuentas", type=["xlsx"])

if st.button("Validar y Descargar"):

    if fecha_desde > fecha_hasta:
        st.error("La fecha inicial no puede ser mayor a la final.")
    elif plantilla_file is None:
        st.error("Debe subir el archivo de plantilla.")
    else:
        fecha_desde_str = fecha_desde.strftime("%Y-%m-%d")
        fecha_hasta_str = fecha_hasta.strftime("%Y-%m-%d")

        st.info("🔍 Consultando cantidad de registros…")
        cantidad = conteo(tipo_entidad,fecha_desde_str, fecha_hasta_str)
        st.write(f"📌 Registros: **{cantidad:,}**")

        st.info("⬇️ Descargando datos…")
        df = descargar_datos(tipo_entidad,fecha_desde_str, fecha_hasta_str)

        st.info("🔧 Procesando…")
        pivot_df = procesar_dataframe(df, plantilla_file, factor_rango)

        st.success("Archivo generado correctamente.")

        xlsx_bytes = generar_excel(
            pivot_df,
            tipo_entidad,
            fecha_desde_str,
            factor_rango,
            etiqueta_unidad
        )

        st.download_button(
            label="📥 Descargar archivo XLSX",
            data=xlsx_bytes,
            file_name=f"{nombre_hoja[tipo_entidad]}{fecha_desde.strftime('%m%Y')}n.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
